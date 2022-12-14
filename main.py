from openpyxl import load_workbook
import json

def get_json_values() -> dict[str, str]:
    try:
        with open('./app-info.json', 'r') as json_file:
            data = json.load(json_file)
            return {
                "EXCEL_FILE_NAME": data["EXCEL_FILE_NAME"],
                "CSV_FILE_NAME": data["CSV_FILE_NAME"],
                "SQL_FILE_NAME": data["SQL_FILE_NAME"],
                "SQL_TABLE_NAME": data["SQL_TABLE_NAME"]
            }
    except Exception:
        raise

def get_csv_excel(excel_file_name: str, csv_file_name: str) -> bool:
    wb = load_workbook(filename=f"{excel_file_name}")
    ws = wb.active
    try:
        with open(f"{csv_file_name}", 'w') as text_file:
            for row in ws.iter_rows(min_row=2):
                counter = 1
                for cell in row:
                    text_file.write(f"{cell.value},") if counter < ws.max_column else text_file.write(f"{cell.value}\n")
                    counter += 1
    except Exception:
        raise
    finally:
        wb.close()
    return True


def get_sql_query(excel_file_name: str, csv_file_name: str, sql_file_name: str, sql_table_name: str) -> str:

    def get_column_name_excel() -> list[str]:
        wb = load_workbook(filename=f"{excel_file_name}")
        ws = wb.active
        try:
            return [cell.value 
                    for row in ws.iter_rows(max_row=1)
                    for cell in row]
        except Exception:
            raise
        finally:
            wb.close()


    def get_column_name_sql() -> str:
        list_column_names_excel = get_column_name_excel()
        fields_insert_value = f"INSERT INTO {sql_table_name} ("
        counter = 0
        for field in list_column_names_excel:
            if counter < len(list_column_names_excel) - 1:
                fields_insert_value += f'{field}, '
            else:
                fields_insert_value += f'{field}) VALUES \n'
            counter += 1
        return fields_insert_value


    def get_insert_values_sql() -> bool:
        column_headers_name_sql = get_column_name_sql()
        try:
            with open(f'./{csv_file_name}', 'r') as read_csv_file:
                with open(f'./{sql_file_name}', 'w') as write_sql_file:
                    write_sql_file.write(column_headers_name_sql)
                    for row in read_csv_file:
                        row = row.replace("\n", "")
                        counter = 0
                        write_sql_file.write('(')
                        for field in row.split(','):
                            row_size = len(row.split(','))
                            
                            if counter < row_size - 1:
                                write_sql_file.write(f"'{field}', ")

                            if counter == row_size - 1:
                                write_sql_file.write(f"'{field}'")

                            counter += 1
                        write_sql_file.write(f"),\n")
        except Exception:
            raise
        return True

    return get_insert_values_sql()    


def main():    
    json_values = get_json_values()

    EXCEL_FILE_NAME = json_values['EXCEL_FILE_NAME']
    CSV_FILE_NAME   = json_values['CSV_FILE_NAME']
    SQL_FILE_NAME   = json_values['SQL_FILE_NAME']
    SQL_TABLE_NAME  = json_values['SQL_TABLE_NAME']

    if get_csv_excel(EXCEL_FILE_NAME, CSV_FILE_NAME):
        print(f'{CSV_FILE_NAME} GENERATED.')
        
        if get_sql_query(EXCEL_FILE_NAME, CSV_FILE_NAME, SQL_FILE_NAME, SQL_TABLE_NAME):
            print(f'{SQL_FILE_NAME} GENERATED.')

if __name__ == '__main__':
    main()